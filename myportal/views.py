from email import message
import re
import smtplib
from myportal.function.functions import handle_uploaded_file
from django.contrib import messages
from django.http.response import HttpResponse
from email.message import EmailMessage
from django.shortcuts import render, get_object_or_404, redirect, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
import pyexcel as p
import xls2xlsx
from .forms import FileUploadForm, SearchUIDForm, MailContentForm, AddDberForm
import xlrd
import openpyxl
from django.core.mail import send_mail
from openpyxl.workbook import Workbook
from xls2xlsx import XLS2XLSX
from openpyxl.reader.excel import load_workbook
from django.conf import settings
from users.models import Profile
from smtplib import SMTP
from .models import *
from users.models import Profile
from users.forms import *

def home(request):
    # for j in Profile.objects.get(user = request.user).count():
    print(request.user)
    user = Profile.objects.filter(user = request.user,qualifier = 'dber')
    
    return render(request, 'myportal/home.html',{'user':user})

@login_required
def upload_file(request):
    
   
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            print("x")
            form.save()
            
            file = form.save(commit=False)
            file.save()
            print('b')
            document = file.doc
            print('a')
           
            
            wb = xlrd.open_workbook(document.path)
            print('c')
            sheet = wb.sheet_by_index (0)
            sheet.cell_value(0, 0)

            rows = sheet.nrows
            
            print ("b")
            for i in range(rows-1):
                print('v')
                dber_uid = int(sheet.cell_value(i+1, 0))
                print(dber_uid)
                dber_name = sheet.cell_value(i+1, 1)
                print(sheet.cell_value(i+1, 2))
                
                dber_gender = sheet.cell_value(i+1, 3)
                dber_city = sheet.cell_value(i+1, 4)
                dber_state = sheet.cell_value(i+1, 5)
                print('v')
                Profile.objects.create(user = request.user,name=dber_name,uid=dber_uid,city = dber_city,state = dber_state,gender = dber_gender)
               
            print ("c")
            return redirect('home')
        else:
            print("a")
    else:
        form = FileUploadForm()

        return render(request, 'myportal/upload_file.html', {
            'form': form
        })

def search_dber(request):

    if request.method == 'POST':
        form = SearchUIDForm(request.POST)
        if form.is_valid():
            search_uid = form.cleaned_data['uid']
            print('a')
            print(request.user)
            file = get_object_or_404(Portal, staff = Profile.objects.filter(user = request.user)[0])
            document = file.doc
            wb = xlrd.open_workbook(document.path)
            sheet = wb.sheet_by_index(0)
            print('b')
            rows = sheet.nrows
            cols = sheet.ncols
            print('c')
            flag = 0
            print('d')
            for i in range(rows-1):
                print('e')
                if search_uid == sheet.cell_value(i+1, 0):
                    print('d')
                    flag = 1
                    dber = get_object_or_404(Profile, uid = search_uid)
                    return render(request, "myportal/search_results.html", {'dber':dber})

            if flag == 0:
                messages.success(request,'user does not exist')
                return redirect('home')
    else:
        form = SearchUIDForm()

        return render(request, 'myportal/search_dber.html', {
            'form': form
        })




@login_required
def add_dber(request):

    if request.method == 'POST':
        print('a')
        form = AddDberForm(request.POST)
        print('b')
        if form.is_valid():
            print('c')
            form.save()
            dber_uid = form.cleaned_data.get('uid')
            
            
            name = form.cleaned_data.get('name')
            print('x')
            dob = form.cleaned_data.get('dob')
            gender = form.cleaned_data.get('gender')
            city = form.cleaned_data.get('city')
            state = form.cleaned_data.get('state')
            mail = form.cleaned_data.get('mail')
            print('d')
            
            file = get_object_or_404(Portal,  staff = Profile.objects.filter(user = request.user)[0])
            document = file.doc
            print('e')
            
            print(document.path)
            print(document.name)
            p.save_book_as(file_name =document.path,dest_file_name="new.xlsx")
            print('d')
        
            wb = openpyxl.load_workbook('new.xlsx')
            print('f')
            print(wb)
            sheet = wb.active

            rows = sheet.max_row
            cols = sheet.max_column
            print('g')
            sheet.insert_rows(rows+1) 

            list = [dber_uid,name, dob,gender,city,state,mail]
            for j in range(1,7):
                print('h')
                cell = sheet.cell(row = rows+1, column = j)
                cell.value = list[j-1]
            print('i')
       
            wb.save(document.path)
            file.save()
            
            print('j')
            return redirect('home')

    else:
        print('k')
        form = AddDberForm()

        return render(request, 'myportal/add_dber.html', {'form':form})

@login_required
def send_mail(request, pk):
    print(pk)
    print(Profile.objects.filter(id=pk))
    dber = get_object_or_404(Profile, pk = pk)
    if request.method == 'POST':
        form = MailContentForm(request.POST)

        if form.is_valid():

            subject = form.cleaned_data['subject']
            content = form.cleaned_data['content']

            mail = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mail.ehlo()
            mail.starttls()
            mail.ehlo()
            mail.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            # message = f'Subject : {subject} \n\n{content}\n\nRegards,\n{request.user}'
            message = EmailMessage()
            message['Subject'] = subject
            message.set_content(content)
            
            message['From'] = settings.EMAIL_HOST_USER
            
            email = dber.mail
            message['To'] = email
            try:
                mail.send_message(message)
            except:
                pass

            mail.close()
            

            
            return redirect('home')
       
    else:
        form = MailContentForm()

    return render(request, 'myportal/mail_content.html', {'form':form})

@login_required
def send_mass_mails(request):

    if request.method == 'POST':
        form = MailContentForm(request.POST)

        if form.is_valid():

            subject = form.cleaned_data['subject']
            content = form.cleaned_data['content']

            mail = smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT)
            mail.ehlo()
            mail.starttls()
            mail.ehlo()
            mail.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)

            message = f'Subject : {subject} \n\n{content}\n\nRegards,\n{request.user}'

     
            user = Profile.objects.filter(user = request.user)
            for user in user:
                if user.qualifier == 'staff':
                    pass
                else:

                    email = user.mail
                    try:
                        mail.sendmail(settings.EMAIL_HOST_USER, email, message)
                    except:
                        pass

            mail.close()
            return redirect('home')

    else:
        form = MailContentForm()

        return render(request, 'myportal/mail_content.html', {'form':form})





